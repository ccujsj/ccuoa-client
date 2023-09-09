from openpyxl import load_workbook
from schemas import StudentSchema
from session import ClintSession



def reverse_dict(dic:dict):
    a = {}
    for k,v in dic.items():
        a.update({v:k})
    return a

def build_field_mapping(header : list):
    d = {}
    idx = 0
    for i in header:
        d.update({idx:i})
        idx+=1
    return d

def load_excel_to_model(excel_filename:str, model:object.__class__,field_filter):
    data = load_workbook(excel_filename,read_only=True).active
    field_map = {}
    idx = 0
    objects = []
    for row in data.iter_rows():
        if idx == 0:
            field_map = reverse_dict(build_field_mapping(map(lambda x: x.value, row)))
            idx += 1
            continue
        else:
            data_dict = {}
            for attr in field_map.keys():
                k = attr
                v = row[field_map[attr]].value
                v = field_filter(k, v)
                data_dict.update({attr: v})

        objects.append(model(**data_dict))
        idx += 1
    return objects

