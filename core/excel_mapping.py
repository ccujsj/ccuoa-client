from openpyxl import Workbook,load_workbook
from client.StudentInfo import StudentSchema, StudentSchema_filter
from client.MoralScore import MoralRecordSchema, MoralRecordSchema_filter
from client.InfoText import TextSchema, TextSchema_filter
from core.algorithms import jaccard_similarity, in_seq_max

schemas = {
    "Student":StudentSchema,
    "Moral":MoralRecordSchema,
    "Text":TextSchema
}


class ExcelModel:
    def __init__(self,filename,auto_load:bool=False):
        # load excel workbook
        wb:Workbook = load_workbook(filename)
        self.__wb = wb
        # confirm a prime data sheet and a mapping sheet
        self.__obj_cls :object.__class__
        self.__mapping_sheet = None  # sheet with mapping
        self.__data_sheet = None
        self.__loaded = False
        self.__mapping = {}
        if not auto_load:
            return
        # by mark position
        mark_pos = 'A1'
        for i in wb.worksheets:
            mark_point:str = i[mark_pos].value
            if mark_point.startswith("#"):
                self.obj_cls:object.__class__ = schemas.get(mark_point[1:])
                self.__mapping_sheet = i
                self.__loaded = True
                break
        if self.__mapping is None:
            return
        for i in self.__mapping_sheet.iter_rows():
            field_name = i[1].value
            class_attr = i[0].value
            self.__mapping.update({class_attr:field_name})


        # ------- Auto Load Process START----------
        """
        This part is not finished yet, but I'm not going to code it any more.
        Auto load process is based on the similarity between head line and mapping.
        Author has no confident about weather it is right, because it used a lot of 
        magic methods and border check is also not being good taken care of.
        To be honest, this automatic loading operation does not meet the standards 
        of software engineering, and it is done because it can speed up efficiency 
        under reasonable administrative supervision, which needs administrative support
                                                      -- Fei Dong xu 2023-9-13
        """
        for i in schemas.items():
            cls_obj:object.__class__ = i[1]
            cls_obj_list = []
            for j in vars(cls_obj).get('__fields__').items():
                obj_attr = j[0]
                cls_obj_list.append(obj_attr)
            ## match part
        headers = []  # get every sheets' head-line (title line) [TITLE LINE LIST]
        for sheet in wb.worksheets:
            sheet_name = sheet.title
            for j in sheet.iter_rows():
                col = list(j)
                col_val = []
                for val in col:
                    col_val.append(val.value)
                headers.append((sheet_name,col_val))
                break

        similarity_seq = []
        for sheet_header in headers:
            head_line=sheet_header[1]
            similarity_seq.append((sheet_header[0],jaccard_similarity(head_line,self.__mapping.values())))
        res_dict = {}
        for item_map in similarity_seq:
            res_dict.update({item_map[1]:item_map[0]})

        result = in_seq_max(list(res_dict.keys()))
        matched_sheet_title = res_dict[result]
        self.__data_sheet = None
        for sheet in wb.worksheets:
            if sheet.title == matched_sheet_title:
                self.__data_sheet = sheet
                break
        self.__loaded = True
        self.__auto_loaded = True
        self.__loaded_data = None
        # ------- Auto Load Process ENDED -------
    def set_mapping(self,map_dict):
        self.__mapping.update(map_dict)
        self.__loaded = True
        self.__data_sheet = self.__wb.active
    def instantiation(self,mapping:dict,class_schema:object.__class__):
        """
        Multiple class instantiations can easily cause problems
        however this func is recommended
        :param mapping:
        :param class_schema:
        :return:
        """

        self.__data_sheet = self.__wb.active
        obj_attrs = list(vars(class_schema).get("__fields__").keys())
        schema_list = []
        header_line_dict = {}  ## map the field and index
        idx = 0
        # get title line first
        for row in self.__data_sheet.iter_rows():
            for cell in list(row):
                header_line_dict.update({cell.value:idx})
                idx += 1
            break
        count = 0
        for row in self.__data_sheet.iter_rows():
            if count == 0:
                # jump over the title line
                count += 1
                continue
            ## do dump
            json_dict = {}
            for attr in obj_attrs:
                ## attr also alias : field
                value_idx = header_line_dict.get(mapping.get(attr))
                val = row[value_idx].value
                if class_schema == MoralRecordSchema:
                    field_filter = MoralRecordSchema_filter
                elif class_schema == StudentSchema:
                    field_filter = StudentSchema_filter
                elif class_schema == TextSchema:
                    field_filter = TextSchema_filter
                else:
                    field_filter = lambda k,v:v
                json_dict.update({attr:field_filter(attr,val)})
            schema_list.append(MoralRecordSchema(**json_dict))
        return schema_list
    def dump_student(self,mapping:dict):
        """
        I am not going to code this.
        :param mapping:
        :return:
        """
        pass

    def dump_moral(self,mapping:dict):
        """
        its actually the same function as instantiation
        :param mapping:
        :return:
        """
        self.__data_sheet = self.__wb.active
        obj_attrs = list(vars(MoralRecordSchema).get("__fields__").keys())  # ['stu_id', 'stu_name',
        moral_schemas = []
        header_line_dict = {}  ## map the field and index
        idx = 0
        for row in self.__data_sheet.iter_rows():
            for cell in list(row):
                header_line_dict.update({cell.value:idx})
                idx += 1
            break
        count = 0
        for row in self.__data_sheet.iter_rows():
            if count == 0:
                count += 1
                continue
            ## do dump
            json_dict = {}
            for attr in obj_attrs:
                ## attr also alias : field
                value_idx = header_line_dict.get(mapping.get(attr))
                val = row[value_idx].value
                json_dict.update({attr:MoralRecordSchema_filter(attr,val)})
            moral_schemas.append(MoralRecordSchema(**json_dict))
        return moral_schemas
        pass
    def dump_text(self,mapping:dict):
        pass

    @property
    def loaded(self) -> bool:
        return self.__loaded


def reverse_dict(dic:dict):  # disabled
    a = {}
    for k,v in dic.items():
        a.update({v:k})
    return a

def build_field_mapping(header : list):  # disabled
    d = {}
    idx = 0
    for i in header:
        d.update({idx:i})
        idx+=1
    return d

def load_excel_to_model(excel_filename:str, model:object.__class__,field_filter):  # disabled
    data = load_workbook(excel_filename,read_only=True).active
    field_map = {}
    idx = 0
    objects = []
    for row in data.iter_rows():
        if idx == 0:
            field_map = reverse_dict(build_field_mapping(map(lambda x: x.value, row)))
            idx += 1
            continue
        else:
            data_dict = {}
            for attr in field_map.keys():
                k = attr
                v = row[field_map[attr]].value
                v = field_filter(k, v)
                data_dict.update({attr: v})

        objects.append(model(**data_dict))
        idx += 1
    return objects


